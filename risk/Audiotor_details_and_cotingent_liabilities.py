import pandas as pd
from datetime import datetime, timedelta
import psycopg2
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
import datetime as dt
# from NBFC_Company_list import NBFC_list, check_nbfc, check_nbfc_name
# from open_ai_prompt_code import get_financial_concentration_insights
import os
import re
import json
# from consolidate_pdf import consolidate_EBIDTA
from datetime import date
import requests
from bs4 import BeautifulSoup
from urllib.parse import urlencode, quote
from lxml import html
from dotenv import load_dotenv
import os
load_dotenv()


file_path = r"D:\\Nexensus_Projects\\risk\\Dalmia Cement (Bharat) Limited-Nexensus Risk Report.xlsx"
conn = psycopg2.connect(
    dbname=os.getenv("DB_NAME"),
    user=os.getenv("DB_USER"),
    password=os.getenv("DB_PASSWORD"),
    host=os.getenv("DB_HOST"),
    port=os.getenv("DB_PORT")
)

# Create a cursor object using the connection
cur = conn.cursor()

############## code for auditor change for updated logic covering missing data logic ###############################################

# company_name = 'TECHNOMED DEVICES INDIA PRIVATE LIMITED'
def Auditor_change(file_path):
    def format_auditor_message(records):
        lines = ["Auditor Change Summary:"]
        for r in records:
            action = "left" if r["status"] == "LEFT" else "joined"
            lines.append(
                f"• FY {r['financial_year']}: {r['auditor_name']} {action} the company"
            )
        return "\n".join(lines)

    def auditor_no_change_message(auditor_set):
        if not auditor_set:
            return "No auditor change detected."

        auditors = sorted(auditor_set)  # for consistent output

        if len(auditors) == 1:
            names = auditors[0]
        elif len(auditors) == 2:
            names = " and ".join(auditors)
        else:
            names = ", ".join(auditors[:-1]) + " and " + auditors[-1]

        return f"No auditor changed from {names}."
    df_synopsis = pd.read_excel(file_path, sheet_name="Company Profile - Synopsis")

    row_index_cin = 5
    col_index_cin = 2
    cin = df_synopsis.iat[row_index_cin, col_index_cin]
    print("cin of company :", cin)
    # cin = 'U65910UP1993PLC015039'

    sql_queries_1 = """
        SELECT auditor_pan, auditor_name, financial_year
        FROM company_core.t_company_auditor_details
        WHERE cin = %s
        AND financial_year BETWEEN
            EXTRACT(YEAR FROM CURRENT_DATE)::int - 4
            AND EXTRACT(YEAR FROM CURRENT_DATE)::int
        ORDER BY financial_year DESC;
    """

    cur.execute(sql_queries_1, (cin,)) 

    # cur.execute(sql_queries_1)
    results = cur.fetchall()
    from collections import defaultdict

    year_to_auditors = defaultdict(dict)
    all_years = set()

    from collections import defaultdict

    year_to_auditors = defaultdict(dict)

    for pan, name, year in results:
        year_to_auditors[year][pan] = name

    sorted_years = sorted(year_to_auditors.keys())

    # print(year_to_auditors)

    events = []

    for i in range(1, len(sorted_years)):
        prev_year = sorted_years[i - 1]
        curr_year = sorted_years[i]

        prev_pans = set(year_to_auditors[prev_year].keys())
        curr_pans = set(year_to_auditors[curr_year].keys())

        year_gap = curr_year - prev_year

        # 🔹 CASE 1: Consecutive years → normal logic
        if year_gap == 1:
            for pan in curr_pans - prev_pans:
                events.append({
                    "financial_year": curr_year,
                    "auditor_pan": pan,
                    "auditor_name": year_to_auditors[curr_year][pan],
                    "status": "JOINED"
                })

            for pan in prev_pans - curr_pans:
                events.append({
                    "financial_year": curr_year,
                    "auditor_pan": pan,
                    "auditor_name": year_to_auditors[prev_year][pan],
                    "status": "LEFT"
                })

        # 🔹 CASE 2: Gap exists
        elif year_gap > 1:
            # SAME auditors → treat as continuation
            if prev_pans == curr_pans:
                continue

            # DIFFERENT auditors → treat as LEFT & JOINED
            for pan in prev_pans - curr_pans:
                events.append({
                    "financial_year": prev_year + 1,  # logical leave year
                    "auditor_pan": pan,
                    "auditor_name": year_to_auditors[prev_year][pan],
                    "status": "LEFT"
                })

            for pan in curr_pans - prev_pans:
                events.append({
                    "financial_year": curr_year,
                    "auditor_pan": pan,
                    "auditor_name": year_to_auditors[curr_year][pan],
                    "status": "JOINED"
                })



    # print(events)

    # if len(set(events)) > 1:
    if len(events) > 1:
        # auditor_changed = True
        # Dynamic message
        rating_msg = f"Change in Auditor"
        message = (
            format_auditor_message(events)
        )
        fill = PatternFill(start_color="ba141a", end_color="ba141a", fill_type="solid")  # red
    else:
        # auditor_changed = False
        # Dynamic message
        rating_msg = f"No Change in Auditor"
        auditor_set = set()
        for d in results:
            auditor_set.add(d[1])
        message = (
            auditor_no_change_message(auditor_set)
        )
        fill = PatternFill(start_color="0c9613", end_color="0c9613", fill_type="solid")  # green


# --- Step 3: Open workbook to update Risk Profile ---
    wb = load_workbook(file_path)
    # print(wb.sheetnames)
    ws = wb["Risk Profile "]
    # auditor_changed = True
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=3).value and "Auditor Change" in str(ws.cell(row=row, column=3).value):

            # --- Column F: Insert message with hyperlink ---
            cell_f = ws.cell(row=row, column=6)
            cell_f.value = message
            # cell_f.hyperlink = f"#'Other Financial Items'!B{target_row}"
            # cell_f.style = "Hyperlink"
            cell_f.alignment = Alignment(wrap_text=True, vertical="top")

            # Ensure gridlines visible
            thin = Side(border_style="thin", color="000000")
            cell_f.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # Adjust width & height
            ws.column_dimensions['F'].width = 90  
            approx_lines = len(message) // 90 + 1
            ws.row_dimensions[row].height = approx_lines * 15

            # --- Column E: Apply color ---
            cell_e = ws.cell(row=row, column=5)
            cell_e.value = None
            cell_e.fill = fill
            cell_e.alignment = Alignment(horizontal="center", vertical="center")
            cell_e.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # --- Column D: Apply Short Message ---
            cell_d = ws.cell(row=row, column=4)
            cell_d.value = rating_msg
            # if auditor_changed:
            #     cell_d.value = f"Change in Auditor"
            # else:
            #     cell_d.value = f"No Change in Auditor"
            cell_d.alignment = Alignment(horizontal="left", vertical="center")
            cell_d.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            print(f"✅ Inserted Auditor change")
            break

    wb.save(file_path)
    print("file_saved")

# # Example usage:
output = Auditor_change(file_path)

############## code for contingent liabilites for updated logic ###############################################

def contigent_liabilites(file_path):
    df_synopsis = pd.read_excel(file_path, sheet_name="Company Profile - Synopsis")

    row_index_cin = 5
    col_index_cin = 2
    cin = df_synopsis.iat[row_index_cin, col_index_cin]

    # cin = 'U65191TN1996PLC035963'
    report_financial_year = 2025
    # ROUND((bs.total_contingent_liabilities::numeric / csaf.reserves_surplus) * 100, 2) AS sp_percentage

    sql_queries_1 = """
        SELECT
        cd.cin,
        cd.company_name,
        bs.total_contingent_liabilities,
        bs.financial_year,
        csaf.reserves_surplus,
        ROUND(
            (
                COALESCE(bs.total_contingent_liabilities, 0)::numeric
                / NULLIF(COALESCE(csaf.reserves_surplus, 0), 0)
            ) * 100,
            2
        ) AS sp_percentage
        FROM company_core.t_company_detail cd
        JOIN company_core.t_company_balance_sheet bs
            ON bs.company_id = cd.id
        JOIN company_core.t_company_source_application_funds csaf
            ON csaf.company_id = cd.id
        AND csaf.financial_year = bs.financial_year
        WHERE cd.cin = %s
        ORDER BY bs.financial_year DESC limit 1;
    """

    cur.execute(sql_queries_1, (cin,)) 

    # cur.execute(sql_queries_1)
    results = cur.fetchall()
    # print("results", results)
    # results = [('L99999MH1946PLC004768', 'Larsen And Toubro Limited', 0, 2025, 600, 0.00)]
    financial_year = results[0][3]
    if financial_year == report_financial_year:
        total_contingent_liabilities = results[0][2]
        reserve_surplus = results[0][4]
        if total_contingent_liabilities is not None:
            total_contingent_liabilities = total_contingent_liabilities / 10_000_000  # Convert Total Contingent Liabilities to Crores
        else:
            total_contingent_liabilities = 0
        CL_percent = results[0][5]
        if CL_percent is None:
            CL_percent = 0 
        if reserve_surplus is not None:
            reserve_surplus = reserve_surplus / 10_000_000  # Convert Total net worth to Crores
        else:
            reserve_surplus = 0
        # print([total_contingent_liabilities, financial_year, reserve_surplus, CL_percent])
        # if CL_percent < 0 or CL_percent > 25:
        if total_contingent_liabilities <= 0 and reserve_surplus <= 0:
            fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid") # yellow
            if total_contingent_liabilities == 0 and reserve_surplus < 0:
                rating_msg = (
                    f"Total contingent liability and commitment~ Nil.\nTotal reserve and surplus is Negative"
                )
                message = (
                f"The Total contingent liability and commitment is nil in FY{str(results[0][3])[2:]} and Total Reserve and surplus is negative INR~({reserve_surplus}) Cr in FY{str(results[0][3])[2:]}."
                )
            elif total_contingent_liabilities < 0 and reserve_surplus < 0:
                rating_msg = (
                    f"Total contingent liability and commitment is Negative.\nTotal reserve and surplus is negative"
                )
                message = (
                f"The Total contingent liability and commitment is negative INR~({total_contingent_liabilities}) Cr's in FY{str(results[0][3])[2:]}  and surplus and Total Reserve and surplus is negative INR~({reserve_surplus}) Cr in FY{str(results[0][3])[2:]}"
                )
            elif total_contingent_liabilities == 0 and reserve_surplus == 0:
                rating_msg = (
                    f"Total contingent liability and commitment ~ Nil.\nTotal reserve and surplus ~ Nil"
                )
                message = (
                f"The Total contingent liability and commitment is nil in FY{str(results[0][3])[2:]} and Total Reserve and surplus is Nil in FY{str(results[0][3])[2:]}"
                )
            elif total_contingent_liabilities < 0 and reserve_surplus == 0:
                rating_msg = (
                    f"T Total contingent liability and commitment is negative.\nTotal reserve and surplus ~ Nil"
                )
                message = (
                f"The Total contingent liability and commitment is negative INR~({total_contingent_liabilities}) Cr's in FY{str(results[0][3])[2:]} and Total Reserve and surplus is Nil in FY{str(results[0][3])[2:]}"
                )

        elif total_contingent_liabilities > 0 and reserve_surplus <= 0:
            fill = PatternFill(start_color="ba141a", end_color="ba141a", fill_type="solid")  # red
            if reserve_surplus < 0:
                rating_msg = (
                    f"High Total contingent liability and commitment is present.\nTotal reserve and surplus is Negative."
                )
                message = (
                f"The Total contingent liability and commitment is INR~{total_contingent_liabilities} Cr's in FY{str(results[0][3])[2:]} and Total Reserve and surplus is negative INR~({reserve_surplus}) Cr in FY{str(results[0][3])[2:]}"
                )
            elif reserve_surplus == 0:
                rating_msg = (
                    f"High Total contingent liability and commitment is present.\nTotal reserve and surplus ~ Nil"
                )
                message = (
                f"The Total contingent liability and commitment is INR~{total_contingent_liabilities} Cr's in FY{str(results[0][3])[2:]} and Total Reserve and surplus is Nil in FY{str(results[0][3])[2:]}"
                )                

        elif total_contingent_liabilities < 0 and reserve_surplus > 0:
            fill = PatternFill(start_color="0c9613", end_color="0c9613", fill_type="solid")  # green
            rating_msg = (
                f"Total contingent liability and commitment is Negative.\nTotal reserve and surplus is INR~{reserve_surplus} Cr's"
            )
            message = (
            f"The Total contingent liability and commitment is negative INR~({total_contingent_liabilities}) Cr's in FY{str(results[0][3])[2:]}  and surplus and Total Reserve and surplus is  INR~{reserve_surplus} Cr in FY{str(results[0][3])[2:]}"
            )       
        # elif total_contingent_liabilities < 0 and reserve_surplus < 0:
        #     fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid") # yellow
        elif total_contingent_liabilities >= 0 and reserve_surplus >= 0:
            if CL_percent > 25:
                fill = PatternFill(start_color="ba141a", end_color="ba141a", fill_type="solid")  # red
                rating_msg = (
                    f"High contingent liabilities and commitments are present.\nTotal reserve and surplus is INR~{reserve_surplus} Cr's."
                )
                message = (
                f"The Total contingent liability and commitment is INR~{total_contingent_liabilities} Cr's in FY{str(results[0][3])[2:]} , Which is more than {CL_percent}% of total reserve and surplus and Total Reserve and surplus is INR~{reserve_surplus} Cr in FY{str(results[0][3])[2:]}."
                )
            elif CL_percent < 5:
                fill = PatternFill(start_color="0c9613", end_color="0c9613", fill_type="solid")  # green
                rating_msg = (
                    f"Low contingent liabilities and commitments are present.\nTotal reserve and surplus is INR~{reserve_surplus} Cr's."
                )
                message = (
                f"The Total contingent liability and commitment is INR~{total_contingent_liabilities} Cr's in FY{str(results[0][3])[2:]} , Which is more than {CL_percent}% of total reserve and surplus and Total Reserve and surplus is INR~{reserve_surplus} Cr in FY{str(results[0][3])[2:]}."
                )
            else:
                fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid") # yellow
                rating_msg = (
                    f"Moderate contingent liabilities and commitments are present.\nTotal reserve and surplus is INR~{reserve_surplus} Cr's."
                )
                message = (
                f"The Total contingent liability and commitment is INR~{total_contingent_liabilities} Cr's in FY{str(results[0][3])[2:]} , Which is more than {CL_percent}% of total reserve and surplus and Total Reserve and surplus is INR~{reserve_surplus} Cr in FY{str(results[0][3])[2:]}."
                )

        # --- Step 3: Open workbook to update Risk Profile ---
        wb = load_workbook(file_path)
        # print(wb.sheetnames)
        ws = wb["Risk Profile "]
        # auditor_changed = True
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=3).value and "Extent of contingent liabilities" in str(ws.cell(row=row, column=3).value):

                # --- Column F: Insert message with hyperlink ---
                cell_f = ws.cell(row=row, column=6)
                cell_f.value = message
                # cell_f.hyperlink = f"#'Other Financial Items'!B{target_row}"
                # cell_f.style = "Hyperlink"
                cell_f.alignment = Alignment(wrap_text=True, vertical="top")

                # Ensure gridlines visible
                thin = Side(border_style="thin", color="000000")
                cell_f.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                # Adjust width & height
                ws.column_dimensions['F'].width = 90  
                approx_lines = len(message) // 90 + 1
                ws.row_dimensions[row].height = approx_lines * 15

                # --- Column E: Apply color ---
                cell_e = ws.cell(row=row, column=5)
                cell_e.value = None
                cell_e.fill = fill
                cell_e.alignment = Alignment(horizontal="center", vertical="center")
                # cell_e.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                # --- Column D: Apply Short Message ---
                cell_d = ws.cell(row=row, column=4)
                cell_d.value = rating_msg
                cell_d.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                # cell_d.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                print(f"✅ Inserted Extent of contingent liabilities ")
                break
        wb.save(file_path)
        print("file_saved")
        # print(cell_d.value)
contigent_liabilites(file_path)
