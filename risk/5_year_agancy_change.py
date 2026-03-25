import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from datetime import datetime
import psycopg2
import logging

# ================= CONFIG =================

LOG_FILE = r"D:\Nexensus_Projects\risk\script_log.txt"

CATEGORY_1 = {"INDIA RATINGS", "CRISIL", "CARE", "ICRA"}
CATEGORY_2 = {"BRICKWORK", "INFOMERICS", "ACUITE"}

DB_CONFIG = {
    "user": "postgres",
    "password": "postgres",
    "host": "localhost",
    "port": "5432",
    "database": "company_uat"
}

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def log(msg):
    print(msg)
    logging.info(msg)

# ================= DATABASE =================

def get_ratings_df(company_name):
    query = """
        SELECT r.rating_date, r.outlook, r.remarks, r.rating_action,
               insto.value AS instrument,
               r.instrument_details,
               rate.value AS rating,
               m.value AS agency
        FROM company_core.t_company_ratings r
        JOIN company_core.t_company_detail d ON d.id = r.company_id
        JOIN company_core.t_master_data m ON r.agency_id = m.id
        JOIN company_core.t_master_data rate ON r.rating_id = rate.id
        JOIN company_core.t_master_data insto ON r.instrument_id = insto.id
        WHERE d.company_name = %s
        ORDER BY r.rating_date;
    """
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute(query, (company_name,))
    rows = cur.fetchall()
    cols = [d[0] for d in cur.description]
    cur.close()
    conn.close()
    return pd.DataFrame(rows, columns=cols)

# ================= HELPERS =================

def normalize(x):
    return str(x).upper().strip()

def safe(x):
    return "" if x is None else str(x)

def is_withdrawn_row(r):
    txt = (
        safe(r["rating_action"]) + " " +
        safe(r["remarks"]) + " " +
        safe(r["outlook"]) + " " +
        safe(r["rating"])
    ).upper()

    return any(k in txt for k in [
        "WITHDRAWN",
        "NOT MEANINGFUL",
        "NOT RATED",
        "NR",
        "ISSUER NOT COOPERATING"
    ])

# ================= TIME WINDOW (NO DATA LOSS) =================

def prepare_time_window(df):
    df = df.copy()
    df["rating_date"] = pd.to_datetime(df["rating_date"], errors="coerce")

    latest_year = df["rating_date"].dropna().dt.year.max()
    start_year = latest_year - 4

    df["rating_year"] = df["rating_date"].dt.year
    df["in_5yr_window"] = df["rating_year"].between(start_year, latest_year)

    return df, start_year, latest_year

# ================= YEAR-WISE AGENCY STATUS =================

def yearly_agency_status(df_5yr):
    yearly = {}

    for year, grp in df_5yr.groupby("rating_year"):
        active = set()
        withdrawn = set()

        grp = grp.sort_values("rating_date")

        for agency, ag_rows in grp.groupby("agency"):
            agency = normalize(agency)
            active.add(agency)

            latest_row = ag_rows.iloc[-1]
            if is_withdrawn_row(latest_row):
                withdrawn.add(agency)

        yearly[year] = {
            "active": active,
            "withdrawn": withdrawn
        }

    return yearly

# ================= ADD / REMOVE =================

def detect_changes(yearly):
    years = sorted(yearly.keys())
    added, removed = set(), set()

    prev = yearly[years[0]]["active"]

    for y in years[1:]:
        curr = yearly[y]["active"]
        added |= (curr - prev)
        removed |= (prev - curr)
        prev = curr

    return added, removed

# ================= FINAL ANALYSIS =================

def analyze(df):
    df, start_year, latest_year = prepare_time_window(df)
    df_5yr = df[df["in_5yr_window"]]
    print(df_5yr)

    yearly = yearly_agency_status(df_5yr)
    added, removed = detect_changes(yearly)
    print(yearly)
    print(added,"====", removed)
    all_active = set()
    all_withdrawn = set()

    for v in yearly.values():
        all_active |= v["active"]
        all_withdrawn |= v["withdrawn"]

    current_active = yearly[latest_year]["active"]

    # -------- ZONE --------
    if len(all_withdrawn) > len(added):
        zone, color = "Negative Agency Movement", "ba141a"
    elif len(added) > len(all_withdrawn):
        zone, color = "Positive Agency Movement", "0c9613"
    else:
        zone, color = "Stable", "0c9613"

    # -------- MESSAGE --------
    lines = [
        f"Rating Agency Assessment ({start_year}–{latest_year}):",
        "",
        f"Currently Active Agencies ({latest_year}): {', '.join(sorted(current_active)) or 'None'}",
        f"Agencies Withdrawn During Period: {', '.join(sorted(all_withdrawn)) or 'None'}",
        f"New Agencies Added: {', '.join(sorted(added)) or 'None'}",
        f"Agencies Discontinued: {', '.join(sorted(removed)) or 'None'}",
        "",
        "Year-wise Movement:"
    ]

    for y in sorted(yearly):
        a = ", ".join(sorted(yearly[y]["active"]))
        w = ", ".join(sorted(yearly[y]["withdrawn"])) or "None"
        lines.append(f"• {y}: Active – {a} | Withdrawn – {w}")

    lines += [
        "",
        "Interpretation:",
        "An agency is considered withdrawn in a year when its latest rating "
        "reflects withdrawal, non-cooperation, or not-rated status. "
        "Earlier active ratings within the same year do not negate withdrawal."
    ]

    return zone, "\n".join(lines), color

# ================= EXCEL UPDATE =================

def update_excel(file_path, message, zone, color):
    wb = load_workbook(file_path)
    ws = wb["Risk Profile"]

    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    thin = Side(border_style="thin", color="000000")

    for r in range(1, ws.max_row + 1):
        if "Changes in rating agency during last 5 years" in str(ws.cell(r, 3).value):
            ws.cell(r, 4).value = zone
            ws.cell(r, 5).fill = fill

            c = ws.cell(r, 6)
            c.value = message
            c.alignment = Alignment(wrap_text=True, vertical="top")

            for col in [4, 5, 6]:
                ws.cell(r, col).border = Border(
                    top=thin, left=thin, right=thin, bottom=thin
                )
            break

    wb.save(file_path)

# ================= MAIN =================

def rating_agency_changes(file_path, company_name):
    log(f"Processing {company_name}")
    df = get_ratings_df(company_name)

    if df.empty:
        log("No rating data found.")
        return

    zone, msg, color = analyze(df)
    update_excel(file_path, msg, zone, color)
    log(f"{company_name} → {zone}")

# ================= RUN =================

if __name__ == "__main__":
    file_path = r"D:\Nexensus_Projects\risk\20 Microns Limited22.xlsx"
    company_name = "20 Microns Limited"
    rating_agency_changes(file_path, company_name)
