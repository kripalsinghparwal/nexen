import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill


def altam_z_score(file_path):
    # --- Step 1: Read Other Financial Items sheet ---
    df_fin = pd.read_excel(file_path, sheet_name="Other Financial Items")
    col_b = df_fin.iloc[:, 1]

    # Find Z-Score row
    z_score_rows = col_b[col_b.astype(str).str.contains("Z-Score", case=False, na=False)]
    if not z_score_rows.empty:
        idx = z_score_rows.index[0]
        z_score_c = float(df_fin.iloc[idx, 2])  # Convert to float
        z_score_d = float(df_fin.iloc[idx, 3])  
        target_row = idx + 2
    else:
        raise ValueError("No 'Z-Score' found in Column B of Other Financial Items sheet.")

    # Find headers (Financial Years)
    header_rows = col_b[col_b.astype(str).str.contains("Financial Years", case=False, na=False)]
    if not header_rows.empty:
        idxhr = header_rows.index[0]
        header_c = str(df_fin.iloc[idxhr, 2])  
        header_d = str(df_fin.iloc[idxhr, 3])  
        fst_yr = header_c[-3:]  
        snd_yr = header_d[-3:]  
    else:
        raise ValueError("No 'Financial Years' header found in Other Financial Items sheet.")

    # --- Step 2: Zone Classification + Color ---
    if z_score_c > 2.99:
        zone = "Safe Zone"
        fill = PatternFill(start_color="0c9613", end_color="0c9613", fill_type="solid")  # light green
    elif 1.81 <= z_score_c <= 2.99:
        zone = "Grey Zone"
        fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")  # light orange
    else:
        zone = "Distress Zone"
        fill = PatternFill(start_color="ba141a", end_color="ba141a", fill_type="solid")  # light red

    if z_score_c > z_score_d:
        meter = "increased"
    else:
        meter = "decreased"
    # --- Step 3: Open workbook to update Risk Profile ---
    wb = load_workbook(file_path)
    ws = wb["Risk Profile"]

    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=3).value and "Altman Z Score" in str(ws.cell(row=row, column=3).value):
            
            # Dynamic message
            message = (
                f"On Alert. This is {zone} and one should exercise Caution. "
                f"Z Score is marginally {meter} from ~{z_score_d:.2f} in FY{snd_yr} "
                f"to ~{z_score_c:.2f} in FY{fst_yr}."
            )

            # --- Column F: Insert message with hyperlink ---
            cell_f = ws.cell(row=row, column=6)
            cell_f.value = message
            cell_f.hyperlink = f"#'Other Financial Items'!B{target_row}"
            cell_f.style = "Hyperlink"
            cell_f.alignment = Alignment(wrap_text=True, vertical="top")

            # Ensure gridlines visible
            thin = Side(border_style="thin", color="000000")
            cell_f.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # Adjust width & height
            ws.column_dimensions['F'].width = 90  
            approx_lines = len(message) // 90 + 1
            ws.row_dimensions[row].height = approx_lines * 15

            # --- Column E: Apply zone color ---
            cell_e = ws.cell(row=row, column=5)
            cell_e.value = None
            cell_e.fill = fill
            cell_e.alignment = Alignment(horizontal="center", vertical="center")
            cell_e.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # --- Column D: Apply Short Message ---
            cell_d = ws.cell(row=row, column=4)
            cell_d.value = f"Z Score is~{z_score_c:.2f} in FY{fst_yr}"
            cell_d.alignment = Alignment(horizontal="center", vertical="center")
            cell_d.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            print(f"✅ Inserted Altman Z Score")
            break

    wb.save(file_path)

def Constitution(file_path):
    df_synopsis = pd.read_excel(file_path, sheet_name="Company Profile - Synopsis")
    row_index_type = 4
    col_index_type = 2
    comp_type = df_synopsis.iat[row_index_type, col_index_type]

    row_index_cin = 5
    col_index_cin = 2
    cin = df_synopsis.iat[row_index_cin, col_index_cin]

    row_index_detail = 6
    col_index_detail = 2
    detail = df_synopsis.iat[row_index_detail, col_index_detail]
    row_index_detail2 = 7
    detail2 = df_synopsis.iat[row_index_detail2, col_index_detail]

    if len(cin) == 21:
        if cin[0] == "L":
            listed_unlisted = "Listed"
            fill = PatternFill(start_color="0c9613", end_color="0c9613", fill_type="solid")  # light green
        elif cin[0] == "U":
            listed_unlisted = "Unlisted"
            fill = PatternFill(start_color="0c9613", end_color="0c9613", fill_type="solid")  # light green
        elif cin[0] == "F":
            listed_unlisted = ""
            fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")  # light green

    reasning_text = f"{listed_unlisted} {detail}; {detail2}"

    nexensus_rating_text = f"{listed_unlisted} {comp_type} Company"
  

    # --- Step 3: Open workbook to update Risk Profile ---
    wb = load_workbook(file_path)
    ws = wb["Risk Profile"]

    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=3).value and "Constitution" in str(ws.cell(row=row, column=3).value):
            

            # --- Column F: Insert message with hyperlink ---
            cell_f = ws.cell(row=row, column=6)
            cell_f.value = reasning_text
            cell_f.hyperlink = f"#'Company Profile - Synopsis'!C{row_index_detail2+1}"
            cell_f.style = "Hyperlink"
            cell_f.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

            # Ensure gridlines visible
            thin = Side(border_style="thin", color="000000")
            cell_f.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # Adjust width & height
            ws.column_dimensions['F'].width = 90  
            approx_lines = len(reasning_text) // 90 + 1
            ws.row_dimensions[row].height = approx_lines * 35

            # --- Column E: Apply zone color ---
            cell_e = ws.cell(row=row, column=5)
            cell_e.value = None
            cell_e.fill = fill
            cell_e.alignment = Alignment(horizontal="center", vertical="center")
            cell_e.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # --- Column D: Apply Short Message ---
            cell_d = ws.cell(row=row, column=4)
            cell_d.value = nexensus_rating_text
            cell_d.alignment = Alignment(horizontal="center", vertical="center")
            cell_d.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            print(f"✅ Inserted Constitution")
            break

    wb.save(file_path)


file_path = r"D:\Projects\Risk Report\Data\Tata Motors Limited.xlsx"
altam_z_score(file_path)
Constitution(file_path)