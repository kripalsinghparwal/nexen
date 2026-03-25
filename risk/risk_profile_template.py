from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import date
import os
import math


def creating_template_for_risk_profile(tata_file, output_file):
    today = date.today().strftime(r"%d-%b %Y")
    comp_name = os.path.basename(tata_file).replace(".xlsx", "").strip().title()

    # Load Tata workbook
    wb = load_workbook(tata_file)

    # Remove existing Risk Profile if present
    if "Risk Profile" in wb.sheetnames:
        wb.remove(wb["Risk Profile"])

    # Create new sheet in 1st position
    ws = wb.create_sheet("Risk Profile", 0)

    # Styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")   # Blue
    header_font = Font(size=12, bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Grey
    section_font = Font(size=11, bold=True)
    normal_font = Font(size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Border styles
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")

    def set_border(cell, row, col, start_row, end_row, start_col, end_col):
        """Apply thin borders inside and medium borders outside table"""
        top = medium if row == start_row else thin
        bottom = medium if row == end_row else thin
        left = medium if col == start_col else thin
        right = medium if col == end_col else thin
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # Title
    ws.merge_cells("B2:F2")
    cell = ws["B2"]
    cell.value = f"Nexensus Risk Report                 {comp_name}                Date: {today}"
    cell.font = Font(size=18, color="FF0000")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    # increase row height
    ws.row_dimensions[2].height = 40

    # Table header
    headers = ["S. No.", "CRD REQUIREMENTS", "Nexensus Rating", "Colour Indication", "Reasoning"]
    for col, text in enumerate(headers, start=2):
        c = ws.cell(row=3, column=col, value=text)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align

    row = 4

    # Sections & items (hardcoded)
    sections = {
        "I. COMPANY SCORES": [
            "Aggregate Nexensus Risk Score",
            "Altman Z Score",
            "Piotroski F Score",
            "Beneish M Score"
        ],
        "II. COMPANY PARAMETERS": [
            "Constitution",
            "Board Quality",
            "Change in Key Management Personal",
            "Board member linked to Bank/NBFC",
            "Product/ Market Concentration",
            "Debt to EBIDTA of the Company vis-a-vis Group Debt to EBIDTA",
            "Debtors beyond 180 days as per ABS for Last 3 years.",
            "Asset Change",
            "Write Off Position",
            "Advance from Customers/ Unbilled Revenue/ Order Book Position",
            "Industry average Contribution",
            "Industry average ICR",
            "Industry average EBIDTA (Excluding net other income)",
            "Qualifications/Emphasis of Matter in Company's ABS",
            "Auditor Change",
            "Extent of contingent liabilities ",
            "No. of litigations pending including NCLT/ NCLAT cases",
            "Default reporting by the Company (applicable for listed entities) ",
            "Promoter holding for last 5 years with pledge percentage",
            "Promoter holding for last 5 years with pledge percentage",
            "Details of funding options exercised by the Unit during last 5 years",
            "Premium/ discount if pricing of funding option. (for funding raised during last 12 months)",
            "EBIDTA Margin - Trend in Movement ",
            "Other Indications - Movement in Cash on Hand / Cash Flow Diversion",
            "Visibility of beneficial ownership",
            "Changes in rating agency during last 5 years",
            "Changes in rating during last 3 years"
        ],
        "III. MARKET TREND (For Listed Companies)": [
            "Changes in analysts recommendation during last 12 months, for listed entities.",
            "Beta of the Scrip and trading volume as percentage to free float.  Also provide the details of scrip in which index it belongs to ",
            "Company P/E v/s Industry average P/E (For Listed Companies)",
            "Average Daily Traded Volume"
        ],
        "IV. CONDUCT OF FACILITIES WITH LENDERS": [
            "Position of UFCE",
            "Details of short term borrowing from lenders other than Banks (non banking channels)",
            "Level of long term debts compared to peer",
            "Comparability of proposed WC limits (both FBWC and NFBWC) to turnover ratio, with peers",
            "Frequency & Extent of borrowing from non-banking channels",
            "Conduct of Term Loan",
            "Holding Security"
        ],
        "V. PROMOTER & GROUP PARAMETERS": [
            "History of Restructuring  OR Initiation of CIRP against any group entity (other than borrowing entity)",
            "Group Details and Financials (Consolidated), Gearing, Liquidity & availability of positive cash flow from operations",
            "List of activities of the Group",
            "Sponsors experience in the Industry of the company",
            "Key Financial Ratios and their trends for group entities (other than borrower entity) ",
            "History of non-fulfilling of capital commitment(s) by the promoter (in the preceding five years)",
            "History of reneging on personal/ third-party guarantee by Promoter/ Director/ Guarantor entities in group entities (in the preceding five years) ",
            "Rating downgrades of group companies during last 12 months.",
            "Extent of related party transactions",
            "New activities proposed to be entered/ or entered during last 12 months by the group",
            "Commitment to ongoing projects/Capex",
            "Commitment to ongoing projects/Capex",
            "Location of beneficial Sponsor ",
            "History of defaults by other borrower group companies",
            "Major proceedings, litigations against Group and Promoters",
            "Extent of negative news in the media on the Promoters",
            "Complexity of Group Ownership Structure",
            "Complexity of Group Entity Structure",
            "Group Guarantees provided as % of Consolidated group networth.",
            "Share of overseas turnover of the Group ",
            "Penalties by Regulators on the Group including the company",
            "Disputes in the Promoter Group/Family etc.",
            "Politically Exposed Status of the Promoter/s",
            "Guarantees from group companies (other than Promoter/ Holding/ Sponsor Co.,) with acceptable financial standing",
            "Network"
        ],
        "VI. REGULATORY COMPLIANCES": [
            "Statutory liabilities as per last ABS",
            "Movement of Statutory Dues over last three years ",
            "Regulatory compliances ",
            "Compliance levels with laws of the land (for the companies with overseas presence)",
            "Provisions made against the estimated liabilities as suggested by Statutory Auditors ",
            "Whether the company complies with Corporate Governance practice",
            "Existence of Mandatory woman director (if applicable)",
            "Existence of Committee to look into sexual harassment issues",
            "Segregation of Chairman and Managing director roles; whether complied with, if applicable",
            "Compliance with regard to minimum public holding",
            "Directorships of Independent Directors; whether within the norms",
            "Compliance with MSMED Act",
            "No. of whistle blower complaints are more than one during the last year",
            "Position of provision for liabilities like gratuity etc. is done on actuarial basis",
            "GST Return",
            "CSR Expenditure Requirement",
            "Compliance of TDS",
            "Compliance with Industry Standards ( ISO, FSSAI , Agmark , Health etc.)"
        ],
        "VII. ACTIVITY & SECTOR": [
            "Major negative news on activity",
            "Stress in the Industry",
            "Susceptibility to regulations ",
            "Industry Life cycle"
        ]
    }

    # Track table bounds
    start_row, start_col = 3, 2
    end_row, end_col = row, 6  # will update dynamically

    # Set row heights
    ws.row_dimensions[3].height = 30  # header row

    def get_required_height(text, base_height=35, max_chars_per_line=50):
        """Calculate row height dynamically based on text length and line breaks"""
        if not text:
            return base_height
        text = str(text)
        # Count explicit line breaks
        line_breaks = text.count("\n") + 1
        # Longest line for wrapping
        longest_line = max(text.split("\n"), key=len)
        wrapped_lines = math.ceil(len(longest_line) / max_chars_per_line)
        # Final number of lines
        lines = max(line_breaks, wrapped_lines)
        return base_height * lines

    # Fill sections
    for sec, items in sections.items():
        # Section row (merge B:F)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        c = ws.cell(row=row, column=2, value=sec)
        c.font = section_font
        c.fill = section_fill
        c.alignment = left_align
        ws.row_dimensions[row].height = 22  # section row height
        end_row = row

        # Borders for section row
        for col in range(2, 7):
            set_border(ws.cell(row=row, column=col), row, col, start_row, None, start_col, None)
        row += 1

        # Items inside each section
        for i, req in enumerate(items, 1):
            ws.cell(row=row, column=2, value=i).font = normal_font
            ws.cell(row=row, column=3, value=req).font = normal_font

            # placeholders for D and F so height adjusts later
            ws.cell(row=row, column=4, value="")
            ws.cell(row=row, column=6, value="")

            # alignment + wrap text
            for col in range(2, 7):
                if col == 3:
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # --- Dynamic row height based on C, D, F ---
            texts = [
                ws.cell(row=row, column=3).value,  # C
                ws.cell(row=row, column=4).value,  # D
                ws.cell(row=row, column=6).value   # F
            ]
            heights = [get_required_height(t) for t in texts]
            ws.row_dimensions[row].height = max(heights)

            end_row = row
            row += 1

    # Apply borders properly (loop again to finalize thick outer edges)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            set_border(ws.cell(row=r, column=c), r, c, start_row, end_row, start_col, end_col)

    # Adjust column widths
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 90

    # Hide gridlines (disable sheet default gridlines for a clean look)
    ws.sheet_view.showGridLines = False

    # Save
    wb.save(output_file)
    print("Risk Profile sheet created cleanly with grey section headers:", output_file)


# Input & output
tata_file = r"D:\Projects\Risk Report\Raw\Tata Motors Limited.xlsx"
output_file = r"D:\Projects\Risk Report\Data\Tata Motors Limited.xlsx"

creating_template_for_risk_profile(tata_file, output_file)